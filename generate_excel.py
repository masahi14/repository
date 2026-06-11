#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
訪問診療 券管理台帳 Excel生成スクリプト
生活保護・介護保険の訪問診療患者管理用Excelファイルを生成します。
"""

import sys
from datetime import date, datetime

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side,
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import FormulaRule
except ImportError:
    print("エラー: openpyxlがインストールされていません。")
    print("インストール方法: pip install openpyxl")
    sys.exit(1)


# ============================================================
# 定数定義
# ============================================================

OUTPUT_FILENAME = "訪問診療_券管理台帳.xlsx"

# カラーコード
COLOR_HEADER_BG = "1F4E79"
COLOR_HEADER_FG = "FFFFFF"

COLOR_MISEI = "FFE0E0"         # 未請求: 薄赤
COLOR_TSUKI = "FFFACD"         # 月遅れ待ち: 薄黄
COLOR_SEIKYU = "E8F5E9"        # 請求済み: 薄緑
COLOR_MIFUCHAKU = "FFE8CC"     # 券未着-継続追跡: 薄橙

COLOR_DASHBOARD_TITLE_BG = "1F4E79"
COLOR_SECTION_BG = "D6E4F0"

# ステータス値
STATUS_MISEI = "未請求"
STATUS_TSUKI = "月遅れ待ち"
STATUS_SEIKYU = "請求済み"
STATUS_MIFUCHAKU = "券未着-継続追跡"

STATUS_LIST = [STATUS_MISEI, STATUS_TSUKI, STATUS_SEIKYU, STATUS_MIFUCHAKU]

# 券種別
VOUCHER_TYPES = ["医療券", "介護券", "医療券+介護券"]

# カラム定義 (列番号, ヘッダー名, 幅)
COLUMNS = [
    (1,  "No",           8),
    (2,  "患者ID",        12),
    (3,  "患者名",        16),
    (4,  "訪問日",        14),
    (5,  "請求対象月",    14),
    (6,  "券種別",        18),
    (7,  "医療券到着日",  16),
    (8,  "介護券到着日",  16),
    (9,  "請求ステータス", 22),
    (10, "請求日",        14),
    (11, "担当者",        12),
    (12, "メモ",          30),
]

# サンプルデータ (患者ID, 患者名, 訪問日, 請求対象月, 券種別, 医療券到着日, 介護券到着日, ステータス, 請求日, 担当者, メモ)
SAMPLE_DATA = [
    ("P001", "山田 太郎", date(2025, 5, 10), "2025年5月", "医療券+介護券",
     date(2025, 5, 20), date(2025, 5, 22), STATUS_SEIKYU,
     date(2025, 5, 25), "田中", "定期訪問"),
    ("P002", "佐藤 花子", date(2025, 5, 12), "2025年5月", "医療券",
     date(2025, 5, 18), None, STATUS_SEIKYU,
     date(2025, 5, 23), "鈴木", ""),
    ("P003", "鈴木 一郎", date(2025, 5, 15), "2025年5月", "介護券",
     None, None, STATUS_MIFUCHAKU,
     None, "田中", "5/30以降も未着。事務所へ確認要"),
    ("P004", "高橋 美子", date(2025, 5, 20), "2025年5月", "医療券+介護券",
     None, None, STATUS_TSUKI,
     None, "佐藤", "6月上旬に券到着予定"),
    ("P005", "渡辺 健二", date(2025, 5, 28), "2025年5月", "医療券",
     None, None, STATUS_MISEI,
     None, "鈴木", "券待ち"),
]


# ============================================================
# スタイルヘルパー関数
# ============================================================

def make_fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


def make_border(style: str = "thin") -> Border:
    side = Side(style=style)
    return Border(left=side, right=side, top=side, bottom=side)


def make_font(bold: bool = False, color: str = "000000", size: int = 11) -> Font:
    return Font(bold=bold, color=color, size=size, name="メイリオ")


def make_alignment(horizontal: str = "center", vertical: str = "center",
                   wrap: bool = False) -> Alignment:
    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap)


def apply_cell_style(cell, fill=None, font=None, alignment=None, border=None):
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border


# ============================================================
# シート1: 管理台帳
# ============================================================

def build_kanri_sheet(wb: Workbook):
    ws = wb.active
    ws.title = "管理台帳"
    ws.sheet_view.showGridLines = True

    # --- ヘッダー行 ---
    header_fill = make_fill(COLOR_HEADER_BG)
    header_font = make_font(bold=True, color=COLOR_HEADER_FG, size=11)
    header_align = make_alignment("center", "center")
    header_border = make_border("medium")

    for col_idx, header_text, col_width in COLUMNS:
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        apply_cell_style(cell,
                         fill=header_fill,
                         font=header_font,
                         alignment=header_align,
                         border=header_border)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # ヘッダー行の高さ
    ws.row_dimensions[1].height = 28

    # ヘッダー行を固定
    ws.freeze_panes = "A2"

    # --- 列幅の微調整 ---
    ws.column_dimensions["A"].width = 6

    # --- データ検証 (ドロップダウン) ---
    # 券種別 (F列: 患者ID追加により1列右へ)
    dv_voucher = DataValidation(
        type="list",
        formula1='"' + ",".join(VOUCHER_TYPES) + '"',
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="入力エラー",
        error="リストから選択してください。",
    )
    dv_voucher.sqref = "F2:F200"
    ws.add_data_validation(dv_voucher)

    # 請求ステータス (I列: 患者ID追加により1列右へ)
    dv_status = DataValidation(
        type="list",
        formula1='"' + ",".join(STATUS_LIST) + '"',
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="入力エラー",
        error="リストから選択してください。",
    )
    dv_status.sqref = "I2:I200"
    ws.add_data_validation(dv_status)

    # --- 条件付き書式 (行の背景色をI列の値で変える) ---
    status_colors = [
        (STATUS_MISEI,     COLOR_MISEI),
        (STATUS_TSUKI,     COLOR_TSUKI),
        (STATUS_SEIKYU,    COLOR_SEIKYU),
        (STATUS_MIFUCHAKU, COLOR_MIFUCHAKU),
    ]

    for status_val, hex_color in status_colors:
        diff_fill = PatternFill(fill_type="solid", fgColor=hex_color)
        rule = FormulaRule(
            formula=[f'$I2="{status_val}"'],
            fill=diff_fill,
            stopIfTrue=False,
        )
        ws.conditional_formatting.add("A2:L200", rule)

    # --- サンプルデータ ---
    thin_border = make_border("thin")
    data_font = make_font(size=10)
    center_align = make_alignment("center", "center")
    left_align = make_alignment("left", "center")

    for row_idx, row_data in enumerate(SAMPLE_DATA, start=2):
        (patient_id, patient, visit_dt, billing_month, voucher_type,
         med_arrive, care_arrive, status, billing_dt, staff, memo) = row_data

        values = [
            row_idx - 1,       # A: No
            patient_id,        # B: 患者ID
            patient,           # C: 患者名
            visit_dt,          # D: 訪問日
            billing_month,     # E: 請求対象月
            voucher_type,      # F: 券種別
            med_arrive,        # G: 医療券到着日
            care_arrive,       # H: 介護券到着日
            status,            # I: 請求ステータス
            billing_dt,        # J: 請求日
            staff,             # K: 担当者
            memo,              # L: メモ
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border

            # 日付列: D(4), G(7), H(8), J(10)
            if col_idx in (4, 7, 8, 10) and isinstance(value, date):
                cell.number_format = "YYYY/MM/DD"
                cell.alignment = center_align
            elif col_idx == 1:
                cell.alignment = center_align
            elif col_idx == 12:
                cell.alignment = left_align
            else:
                cell.alignment = center_align

        ws.row_dimensions[row_idx].height = 20

    # --- 空データ行のボーダー・フォーマット設定 ---
    for row_idx in range(len(SAMPLE_DATA) + 2, 201):
        for col_idx in range(1, 13):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.font = data_font
            if col_idx in (4, 7, 8, 10):
                cell.number_format = "YYYY/MM/DD"
                cell.alignment = center_align
            elif col_idx == 1:
                cell.alignment = center_align
            elif col_idx == 12:
                cell.alignment = left_align
            else:
                cell.alignment = center_align
        ws.row_dimensions[row_idx].height = 20

    return ws


# ============================================================
# シート2: ダッシュボード
# ============================================================

def build_dashboard_sheet(wb: Workbook):
    ws = wb.create_sheet(title="ダッシュボード")
    ws.sheet_view.showGridLines = False

    # 列幅設定
    col_widths = {"A": 4, "B": 28, "C": 18, "D": 4}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # --- タイトル ---
    ws.merge_cells("B1:C1")
    title_cell = ws["B1"]
    title_cell.value = "訪問診療 券管理 ダッシュボード"
    title_cell.fill = make_fill(COLOR_DASHBOARD_TITLE_BG)
    title_cell.font = make_font(bold=True, color=COLOR_HEADER_FG, size=16)
    title_cell.alignment = make_alignment("center", "center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("B2:C2")
    subtitle_cell = ws["B2"]
    subtitle_cell.value = "生活保護・介護保険 訪問診療患者 管理システム"
    subtitle_cell.fill = make_fill("D6E4F0")
    subtitle_cell.font = make_font(size=11, color="1F4E79")
    subtitle_cell.alignment = make_alignment("center", "center")
    ws.row_dimensions[2].height = 22

    # spacer
    ws.row_dimensions[3].height = 12

    # --- サマリーテーブルタイトル ---
    ws.merge_cells("B4:C4")
    section_cell = ws["B4"]
    section_cell.value = "■ 請求ステータス集計"
    section_cell.fill = make_fill(COLOR_SECTION_BG)
    section_cell.font = make_font(bold=True, size=12, color="1F4E79")
    section_cell.alignment = make_alignment("left", "center")
    ws.row_dimensions[4].height = 26

    # サマリーヘッダー
    for col, text in [("B", "ステータス"), ("C", "件数")]:
        cell = ws[f"{col}5"]
        cell.value = text
        cell.fill = make_fill(COLOR_HEADER_BG)
        cell.font = make_font(bold=True, color=COLOR_HEADER_FG, size=11)
        cell.alignment = make_alignment("center", "center")
        cell.border = make_border("medium")
    ws.row_dimensions[5].height = 24

    # サマリーデータ行
    summary_rows = [
        (STATUS_MISEI,     COLOR_MISEI,     f'=COUNTIF(管理台帳!H:H,"{STATUS_MISEI}")'),
        (STATUS_TSUKI,     COLOR_TSUKI,     f'=COUNTIF(管理台帳!H:H,"{STATUS_TSUKI}")'),
        (STATUS_MIFUCHAKU, COLOR_MIFUCHAKU, f'=COUNTIF(管理台帳!H:H,"{STATUS_MIFUCHAKU}")'),
        (STATUS_SEIKYU,    COLOR_SEIKYU,    f'=COUNTIF(管理台帳!H:H,"{STATUS_SEIKYU}")'),
    ]

    for i, (label, color, formula) in enumerate(summary_rows, start=6):
        row = i
        label_cell = ws.cell(row=row, column=2, value=label)
        label_cell.fill = make_fill(color)
        label_cell.font = make_font(size=11)
        label_cell.alignment = make_alignment("left", "center")
        label_cell.border = make_border("thin")

        count_cell = ws.cell(row=row, column=3, value=formula)
        count_cell.fill = make_fill(color)
        count_cell.font = make_font(bold=True, size=13)
        count_cell.alignment = make_alignment("center", "center")
        count_cell.border = make_border("thin")
        ws.row_dimensions[row].height = 26

    # 合計行
    total_row = 6 + len(summary_rows)
    total_label = ws.cell(row=total_row, column=2, value="合計件数")
    total_label.fill = make_fill("2E75B6")
    total_label.font = make_font(bold=True, color="FFFFFF", size=11)
    total_label.alignment = make_alignment("left", "center")
    total_label.border = make_border("medium")

    total_count = ws.cell(row=total_row, column=3,
                          value="=COUNTA(管理台帳!B2:B200)")
    total_count.fill = make_fill("2E75B6")
    total_count.font = make_font(bold=True, color="FFFFFF", size=13)
    total_count.alignment = make_alignment("center", "center")
    total_count.border = make_border("medium")
    ws.row_dimensions[total_row].height = 28

    # spacer
    spacer_row = total_row + 1
    ws.row_dimensions[spacer_row].height = 14

    # --- 使い方セクション ---
    instr_start = total_row + 2
    ws.merge_cells(f"B{instr_start}:C{instr_start}")
    instr_title = ws[f"B{instr_start}"]
    instr_title.value = "■ システムの使い方（簡易版）"
    instr_title.fill = make_fill(COLOR_SECTION_BG)
    instr_title.font = make_font(bold=True, size=12, color="1F4E79")
    instr_title.alignment = make_alignment("left", "center")
    ws.row_dimensions[instr_start].height = 26

    instructions = [
        "① 「管理台帳」シートに患者情報を入力してください。",
        "② 医療券・介護券が届いたら、到着日を入力します。",
        "③ 請求可能になったらステータスを「請求済み」に更新します。",
        "④ 未請求・追跡状況はこのダッシュボードで確認できます。",
        "⑤ 詳しい使い方は「使い方」シートをご参照ください。",
    ]

    for j, text in enumerate(instructions, start=instr_start + 1):
        ws.merge_cells(f"B{j}:C{j}")
        cell = ws[f"B{j}"]
        cell.value = text
        cell.font = make_font(size=10)
        cell.alignment = make_alignment("left", "center")
        if j % 2 == 0:
            cell.fill = make_fill("F0F4F8")
        ws.row_dimensions[j].height = 22

    # 作成日
    last_row = instr_start + len(instructions) + 2
    ws.merge_cells(f"B{last_row}:C{last_row}")
    date_cell = ws[f"B{last_row}"]
    date_cell.value = f"作成日: {datetime.now().strftime('%Y年%m月%d日')}"
    date_cell.font = make_font(size=9, color="888888")
    date_cell.alignment = make_alignment("right", "center")

    return ws


# ============================================================
# シート3: 使い方
# ============================================================

def build_howto_sheet(wb: Workbook):
    ws = wb.create_sheet(title="使い方")
    ws.sheet_view.showGridLines = False

    # 列幅
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 4

    row = 1

    def add_title(text, r):
        ws.merge_cells(f"B{r}:C{r}")
        cell = ws[f"B{r}"]
        cell.value = text
        cell.fill = make_fill(COLOR_DASHBOARD_TITLE_BG)
        cell.font = make_font(bold=True, color=COLOR_HEADER_FG, size=15)
        cell.alignment = make_alignment("center", "center")
        ws.row_dimensions[r].height = 38
        return r + 1

    def add_section(text, r):
        ws.merge_cells(f"B{r}:C{r}")
        cell = ws[f"B{r}"]
        cell.value = text
        cell.fill = make_fill(COLOR_SECTION_BG)
        cell.font = make_font(bold=True, size=12, color="1F4E79")
        cell.alignment = make_alignment("left", "center")
        ws.row_dimensions[r].height = 26
        return r + 1

    def add_row_two_col(label, content, r, label_bold=False, bg_color=None):
        bg = bg_color or "FFFFFF"
        label_cell = ws.cell(row=r, column=2, value=label)
        content_cell = ws.cell(row=r, column=3, value=content)
        for cell in [label_cell, content_cell]:
            cell.fill = make_fill(bg)
            cell.border = make_border("thin")
            cell.alignment = make_alignment("left", "center", wrap=True)
        label_cell.font = make_font(bold=label_bold, size=10)
        content_cell.font = make_font(size=10)
        ws.row_dimensions[r].height = 22
        return r + 1

    def add_spacer(r, height=10):
        ws.row_dimensions[r].height = height
        return r + 1

    # タイトル
    row = add_title("訪問診療 券管理台帳 - 使い方ガイド", row)
    row = add_spacer(row)

    # --- 概要 ---
    row = add_section("【概要】このシステムについて", row)
    ws.merge_cells(f"B{row}:C{row}")
    cell = ws[f"B{row}"]
    cell.value = ("このExcelファイルは、生活保護・介護保険の訪問診療患者の"
                  "医療券・介護券の管理と請求追跡を効率化するためのツールです。")
    cell.font = make_font(size=10)
    cell.alignment = make_alignment("left", "center", wrap=True)
    ws.row_dimensions[row].height = 36
    row += 1
    row = add_spacer(row)

    # --- 基本的な使い方 ---
    row = add_section("【基本的な使い方】ワークフロー", row)

    steps = [
        ("① 患者情報の登録",
         "訪問実施後すぐに「管理台帳」シートへ患者名・訪問日・請求対象月・券種別・担当者を入力してください。"),
        ("② 初期ステータス設定",
         "券がまだ届いていない場合は請求ステータスを「未請求」に設定します。"),
        ("③ 券到着時の更新",
         "医療券・介護券が届いたら、各到着日欄に日付を入力します（形式: YYYY/MM/DD）。"),
        ("④ 月遅れ対応",
         "当月中に券が届かない場合は「月遅れ待ち」に変更し、翌月以降も追跡します。"),
        ("⑤ 請求完了時",
         "請求が完了したら「請求済み」に変更し、請求日を入力してください。"),
        ("⑥ 未着が長期化した場合",
         "券がなかなか届かない場合は「券未着-継続追跡」に設定し、メモ欄に状況を記録します。"),
        ("⑦ 定期確認",
         "「ダッシュボード」シートで未請求・追跡中の件数を定期的に確認してください。"),
    ]

    for i, (label, content) in enumerate(steps):
        bg = "F8F8FF" if i % 2 == 0 else "FFFFFF"
        row = add_row_two_col(label, content, row, label_bold=True, bg_color=bg)

    row = add_spacer(row)

    # --- ステータスの説明 ---
    row = add_section("【ステータス説明】各ステータスの意味", row)

    # ヘッダー
    for col, text in [(2, "ステータス"), (3, "説明・対応方法")]:
        cell = ws.cell(row=row, column=col, value=text)
        cell.fill = make_fill(COLOR_HEADER_BG)
        cell.font = make_font(bold=True, color=COLOR_HEADER_FG, size=10)
        cell.alignment = make_alignment("center", "center")
        cell.border = make_border("medium")
    ws.row_dimensions[row].height = 22
    row += 1

    status_descriptions = [
        (STATUS_MISEI,    COLOR_MISEI,
         "券がまだ到着していないか、まだ請求手続きが完了していない状態。優先的に確認が必要です。"),
        (STATUS_TSUKI,    COLOR_TSUKI,
         "当月内での請求が間に合わず、翌月以降の月遅れ請求を待っている状態。期限管理に注意してください。"),
        (STATUS_SEIKYU,   COLOR_SEIKYU,
         "券の確認が取れ、請求手続きが完了した状態。請求日を必ず入力してください。"),
        (STATUS_MIFUCHAKU, COLOR_MIFUCHAKU,
         "長期間にわたり券が届いていない状態。関係機関へ問い合わせ・督促が必要です。メモ欄に対応状況を記録してください。"),
    ]

    for status, color, desc in status_descriptions:
        label_cell = ws.cell(row=row, column=2, value=status)
        desc_cell = ws.cell(row=row, column=3, value=desc)
        for cell in [label_cell, desc_cell]:
            cell.fill = make_fill(color)
            cell.font = make_font(size=10)
            cell.border = make_border("thin")
            cell.alignment = make_alignment("left", "center", wrap=True)
        ws.row_dimensions[row].height = 36
        row += 1

    row = add_spacer(row)

    # --- 各列の説明 ---
    row = add_section("【列の説明】管理台帳の各列について", row)

    col_descriptions = [
        ("No",           "自動連番。手動で入力してください。"),
        ("患者名",        "患者のフルネームを入力します。"),
        ("訪問日",        "実際に訪問した日付を入力します（YYYY/MM/DD形式）。"),
        ("請求対象月",    "請求の対象となる月を「YYYY年M月」の形式で入力します（例: 2025年5月）。"),
        ("券種別",        "ドロップダウンから選択: 医療券 / 介護券 / 医療券+介護券"),
        ("医療券到着日",  "医療券が届いた日付。届いていない場合は空欄のままにします。"),
        ("介護券到着日",  "介護券が届いた日付。届いていない場合は空欄のままにします。"),
        ("請求ステータス", "ドロップダウンから選択。行の背景色が自動的に変わります。"),
        ("請求日",        "実際に請求手続きを行った日付。請求済みの場合に入力します。"),
        ("担当者",        "この患者を担当するスタッフ名を入力します。"),
        ("メモ",          "追跡状況や特記事項を自由に記入します。"),
    ]

    for i, (label, content) in enumerate(col_descriptions):
        bg = "F0F4F8" if i % 2 == 0 else "FFFFFF"
        row = add_row_two_col(label, content, row, label_bold=True, bg_color=bg)

    row = add_spacer(row)

    # --- 注意事項 ---
    row = add_section("【注意事項】", row)
    notes = [
        "・データは行2〜200まで入力できます。それ以上必要な場合は行をコピーして追加してください。",
        "・日付のセルはYYYY/MM/DD形式で入力してください。",
        "・ドロップダウン以外の値を入力するとエラーが表示される場合があります。",
        "・ダッシュボードのカウントはCOUNTIF関数を使用しており、自動更新されます。",
        "・定期的にファイルのバックアップを取ることを推奨します。",
    ]
    for note in notes:
        ws.merge_cells(f"B{row}:C{row}")
        cell = ws[f"B{row}"]
        cell.value = note
        cell.font = make_font(size=10)
        cell.alignment = make_alignment("left", "center")
        ws.row_dimensions[row].height = 20
        row += 1

    return ws


# ============================================================
# メイン処理
# ============================================================

def main():
    print("訪問診療 券管理台帳 Excel生成を開始します...")

    try:
        wb = Workbook()

        print("  シート1「管理台帳」を作成中...")
        build_kanri_sheet(wb)

        print("  シート2「ダッシュボード」を作成中...")
        build_dashboard_sheet(wb)

        print("  シート3「使い方」を作成中...")
        build_howto_sheet(wb)

        # 管理台帳を最初のシートとしてアクティブに設定
        wb.active = wb["管理台帳"]

        print(f"  ファイルを保存中: {OUTPUT_FILENAME}")
        wb.save(OUTPUT_FILENAME)

        import os
        abs_path = os.path.abspath(OUTPUT_FILENAME)

        print()
        print("=" * 60)
        print("生成完了!")
        print(f"  ファイル名 : {OUTPUT_FILENAME}")
        print(f"  保存場所   : {abs_path}")
        print()
        print("  含まれるシート:")
        print("    1. 管理台帳      - 患者・券情報の入力・追跡")
        print("    2. ダッシュボード - 請求状況の集計表示")
        print("    3. 使い方        - 操作マニュアル")
        print()
        print("  サンプルデータ: 5件（異なるステータスのサンプルを含む）")
        print("=" * 60)

    except PermissionError:
        print(f"エラー: ファイル '{OUTPUT_FILENAME}' が別のプログラムで開かれています。")
        print("Excelを閉じてから再実行してください。")
        sys.exit(1)
    except Exception as e:
        print(f"エラーが発生しました: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
